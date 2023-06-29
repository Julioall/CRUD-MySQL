from openpyxl import load_workbook

# Função para criar um novo registro
def create_record(file_path, record):
    # Carrega o arquivo Excel
    wb = load_workbook(filename=file_path)
    # Seleciona a planilha ativa
    sheet = wb.active
    # Adiciona o registro como uma nova linha na planilha
    sheet.append(record)
    # Salva as alterações no arquivo
    wb.save(file_path)

# Função para ler todos os registros
def read_records(file_path):
    # Carrega o arquivo Excel
    wb = load_workbook(filename=file_path)
    # Seleciona a planilha ativa
    sheet = wb.active
    # Lê todos os registros na planilha
    records = []
    for row in sheet.iter_rows(values_only=True):
        records.append(row)
    return records

# Função para atualizar um registro existente
def update_record(file_path, record, updated_record):
    # Carrega o arquivo Excel
    wb = load_workbook(filename=file_path)
    # Seleciona a planilha ativa
    sheet = wb.active
    # Procura pelo registro e atualiza os valores correspondentes
    for row in sheet.iter_rows(values_only=True):
        if row == record:
            for index, value in enumerate(updated_record):
                sheet.cell(row=row[0].row, column=index + 1).value = value
            break
    # Salva as alterações no arquivo
    wb.save(file_path)

# Função para excluir um registro existente
def delete_record(file_path, record):
    # Carrega o arquivo Excel
    wb = load_workbook(filename=file_path)
    # Seleciona a planilha ativa
    sheet = wb.active
    # Procura pelo registro e remove a linha correspondente
    for row in sheet.iter_rows(values_only=True):
        if row == record:
            sheet.delete_rows(row[0].row)
            break
    # Salva as alterações no arquivo
    wb.save(file_path)




    
